from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(bold=True, size=14, color="1F4E79")
FONT_SUBTITLE = Font(bold=True, size=11, color="1F4E79")

FILL_ROJO = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
FILL_AMBAR = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
FILL_VERDE = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _estilo_encabezado(ws, ncols):
    for col in range(1, ncols + 1):
        celda = ws.cell(row=1, column=col)
        celda.font = FONT_HEADER
        celda.fill = FILL_HEADER
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = BORDER_THIN


def _auto_ancho(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for celda in col_cells:
            if celda.value is not None:
                max_len = max(max_len, len(str(celda.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def generar_reporte(df_detalle, df_errores, metricas, rates, output_path):
    wb = Workbook()

    # ---------------------------------------------------------------
    # HOJA 1: Resumen Ejecutivo
    # ---------------------------------------------------------------
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ejecutivo"
    ws_resumen.cell(row=1, column=1, value="RESUMEN EJECUTIVO").font = FONT_TITLE
    ws_resumen.merge_cells("A1:F1")

    hoy = date.today()
    ws_resumen.cell(row=3, column=1, value="Fecha de generación:").font = FONT_SUBTITLE
    ws_resumen.cell(row=3, column=2, value=hoy.strftime("%d/%m/%Y"))

    ws_resumen.cell(row=5, column=1, value="MÉTRICAS PRINCIPALES").font = FONT_SUBTITLE
    ws_resumen.cell(row=6, column=1, value="Total CxC (USD)")
    ws_resumen.cell(row=6, column=2, value=metricas["total_cxc_usd"])
    ws_resumen.cell(row=6, column=2).number_format = "#,##0.00"
    ws_resumen.cell(row=7, column=1, value="Total CxP (USD)")
    ws_resumen.cell(row=7, column=2, value=metricas["total_cxp_usd"])
    ws_resumen.cell(row=7, column=2).number_format = "#,##0.00"
    ws_resumen.cell(row=8, column=1, value="Flujo de caja neto (USD)")
    ws_resumen.cell(row=8, column=2, value=metricas["flujo_neto_usd"])
    ws_resumen.cell(row=8, column=2).number_format = "#,##0.00"

    # top 5 vencidas
    ws_resumen.cell(row=10, column=1, value="TOP 5 FACTURAS VENCIDAS").font = FONT_SUBTITLE
    vencidas = (
        df_detalle[df_detalle["dias_vencido"] > 0]
        .sort_values("dias_vencido", ascending=False)
        .head(5)
    )
    headers_top5 = ["Cliente/Proveedor", "Monto USD", "Días vencido"]
    for i, h in enumerate(headers_top5, 1):
        ws_resumen.cell(row=11, column=i, value=h)
    for i, (_, row) in enumerate(vencidas.iterrows()):
        ws_resumen.cell(row=12 + i, column=1, value=row["cliente_proveedor"])
        ws_resumen.cell(row=12 + i, column=2, value=row["monto_usd"])
        ws_resumen.cell(row=12 + i, column=2).number_format = "#,##0.00"
        ws_resumen.cell(row=12 + i, column=3, value=row["dias_vencido"])
    _estilo_encabezado(ws_resumen, 3)

    # tasas usadas
    ws_resumen.cell(row=19, column=1, value="TASAS DE CAMBIO USADAS").font = FONT_SUBTITLE
    ws_resumen.cell(row=20, column=1, value="Moneda")
    ws_resumen.cell(row=20, column=2, value="Tasa (1 USD = ?)")
    _estilo_encabezado(ws_resumen, 2)
    fila_tasa = 21
    for moneda, tasa in sorted(rates.items()):
        if moneda != "USD":
            ws_resumen.cell(row=fila_tasa, column=1, value=moneda)
            ws_resumen.cell(row=fila_tasa, column=2, value=round(tasa, 4))
            fila_tasa += 1

    _auto_ancho(ws_resumen)

    # ---------------------------------------------------------------
    # HOJA 2: Detalle
    # ---------------------------------------------------------------
    ws_detalle = wb.create_sheet("Detalle")
    cols_detalle = [
        "id", "tipo", "cliente_proveedor", "moneda", "monto",
        "fecha_emision", "fecha_vencimiento", "estado",
        "monto_usd", "dias_vencido", "aging",
    ]
    for i, c in enumerate(cols_detalle, 1):
        ws_detalle.cell(row=1, column=i, value=c)
    _estilo_encabezado(ws_detalle, len(cols_detalle))

    for i, (_, row) in enumerate(df_detalle.iterrows()):
        for j, c in enumerate(cols_detalle, 1):
            val = row.get(c)
            ws_detalle.cell(row=2 + i, column=j, value=val)
            celda = ws_detalle.cell(row=2 + i, column=j)
            celda.border = BORDER_THIN

    # formato condicional por aging
    for i, (_, row) in enumerate(df_detalle.iterrows()):
        aging = str(row.get("aging", ""))
        fill_aging = None
        if aging in ("61-90 días", "+90 días"):
            fill_aging = FILL_ROJO
        elif aging == "31-60 días":
            fill_aging = FILL_AMBAR
        elif aging == "Al día":
            fill_aging = FILL_VERDE
        if fill_aging:
            ws_detalle.cell(row=2 + i, column=cols_detalle.index("aging") + 1).fill = (
                fill_aging
            )

    _auto_ancho(ws_detalle)

    # ---------------------------------------------------------------
    # HOJA 3: Inconsistencias
    # ---------------------------------------------------------------
    ws_err = wb.create_sheet("Inconsistencias")
    if not df_errores.empty:
        cols_err = ["fila", "campo", "motivo"]
        for i, c in enumerate(cols_err, 1):
            ws_err.cell(row=1, column=i, value=c)
        _estilo_encabezado(ws_err, len(cols_err))
        for i, (_, row) in enumerate(df_errores.iterrows()):
            for j, c in enumerate(cols_err, 1):
                ws_err.cell(row=2 + i, column=j, value=row[c])
                ws_err.cell(row=2 + i, column=j).border = BORDER_THIN
    else:
        ws_err.cell(row=1, column=1, value="No se encontraron inconsistencias.")

    _auto_ancho(ws_err)

    wb.save(output_path)
